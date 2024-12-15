from datetime import datetime

import openpyxl
from django.core.management.base import BaseCommand
from rich import print
from rich.console import Console
from rich.prompt import Prompt

from gelfcore.logger import log
from gelfmp.management.commands.collectors.types import Aliases
from gelfmp.models import CharcoalMonthlyPlan, Supplier
from gelfmp.utils.normalization import normalize_text_upper

from .collectors.utils import get_best_matches

CHARCOAL_SCHEDULE_PATH = r'.sample-data/schedule_2024.xlsx'

MIN_COL = 2  # onde começa a tabela
MAX_COL = 19  # onde termina a tabela
MIN_ROW = 7  # onde começa a tabela
MONTH_DESLOCATE = 3  # Meses na planilha começam na quarta coluna (índice 3).

NAME_COL = 0
FARM_COL = 1

MINIMUM_SIMILARITY = 95

console = Console()


class Command(BaseCommand):
    def add_arguments(self, parser):
        parser.add_argument('-m', type=int, help='Mês (1-12)', required=True)
        parser.add_argument('-y', type=int, help='Ano', default=datetime.now().year)

    def handle(self, *args, **kwargs):
        month = kwargs['m']
        year = kwargs['y']

        try:
            worksheet = load_schedule(CHARCOAL_SCHEDULE_PATH)
            headers, data_rows = extract_data_rows(worksheet)
            aliases = Aliases('schedule.alias.json')

            validate_month_and_year(headers, month, year)
            save_gelf_schedule(worksheet, month, year)

            process_rows(data_rows, aliases, month, year)

        except Exception as e:
            log.error(f'Erro inesperado durante a coleta de programação de carvão: {e}')


def save_gelf_schedule(worksheet, month, year):
    # pegar a primeira programação pois a primeira, da gelf,
    # tem 3 linhas, uma linha de preço a mais que as outras não tem

    rows = worksheet.iter_rows(values_only=True, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW)

    price_row = next(rows)
    planned_volume_row = next(rows)

    planned_volume = round(planned_volume_row[month + MONTH_DESLOCATE] or 0, 1)
    price = price_row[month + MONTH_DESLOCATE]

    gelf = Supplier.objects.get(corporate_name__icontains='gelf')

    CharcoalMonthlyPlan.objects.update_or_create(
        supplier=gelf,
        month=month,
        year=year,
        defaults={
            'planned_volume': planned_volume,
            'price': price,
        },
    )


def load_schedule(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    worksheet = workbook.active

    if not worksheet:
        raise ValueError(f'Planilha `{path}` não possui uma aba ativa.')

    return worksheet


def extract_data_rows(worksheet):
    headers = next(worksheet.iter_rows(values_only=True, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW - 1))
    data_rows = worksheet.iter_rows(values_only=True, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW + 3)
    return headers, data_rows


def process_rows(data_rows, aliases, month, year):
    suppliers_names = list(Supplier.objects.values_list('corporate_name', flat=True))
    if not suppliers_names:
        raise ValueError('Não existem fornecedores no banco de dados.')

    for row in data_rows:
        # Para a iteração quando chegar ao final da tabela
        if not row[NAME_COL] or row[NAME_COL].lower().strip() == 'total/mês':
            break

        # Pula as linhas de `Realizado` pegando apenas
        # o volume programado para cada fornecedor
        if row[3].lower().strip() != 'programado':
            continue

        schedule_supplier_name = normalize_text_upper(row[NAME_COL] + ' ' + row[FARM_COL])

        # Verifica se o nome já foi mapeado para
        # um fornecedor existente no banco de dados
        if schedule_supplier_name in aliases:
            process_data(aliases.get(schedule_supplier_name), row, month, year)
            continue

        best_matches = get_best_matches(schedule_supplier_name, suppliers_names)
        best_similarity, supplier_name = best_matches[0]

        if best_similarity >= MINIMUM_SIMILARITY:
            process_data(supplier_name, row, month, year)
            continue

        handle_supplier_resolution(
            schedule_supplier_name,
            best_matches,
            aliases,
            row,
            month,
            year,
        )


def process_data(supplier_name, row, month, year):
    try:
        planned_volume = float(row[month + MONTH_DESLOCATE]) if row[month + MONTH_DESLOCATE] else 0

    except ValueError:
        print(f'Volume inválido para o fornecedor {row[NAME_COL]} na linha {row}.')

    supplier = Supplier.objects.get(corporate_name=supplier_name)

    CharcoalMonthlyPlan.objects.update_or_create(
        supplier=supplier,
        month=month,
        year=year,
        defaults={
            'planned_volume': planned_volume,
        },
    )


def handle_supplier_resolution(schedule_supplier_name, best_matches, aliases, row, month, year):
    console.print(f'\n[bold red]FORNECEDOR NÃO ENCONTRADO:[/bold red] {schedule_supplier_name}')
    console.print('\n[bold yellow]Escolha uma das opções abaixo ou pressione Enter para pular.[/bold yellow]\n')

    for i, (similarity, name) in enumerate(best_matches, start=1):
        console.print(f'[cyan]{i}.[/cyan] {name} [dim](Similaridade: {similarity:.2f}%)[/dim]')

    option = Prompt.ask('\n[green]Digite o número da opção escolhida[/green]', default='').strip()

    if option.isdigit():
        option = int(option) - 1
        if 0 <= option < len(best_matches):
            supplier_name = best_matches[option][1]
            aliases.add(schedule_supplier_name, supplier_name)
            process_data(supplier_name, row, month, year)
            console.print(
                f'[bold green]Fornecedor mapeado com sucesso![/bold green]{schedule_supplier_name} -> {supplier_name}\n'
            )
        else:
            console.print('[bold yellow]Opção inválida. Nenhuma ação foi realizada.[/bold yellow]\n')
    else:
        console.print('[bold yellow]Nenhuma correspondência selecionada. Continuando...[/bold yellow]\n')


def validate_month_and_year(headers, month, year):
    try:
        sheet_date = headers[month + MONTH_DESLOCATE]

        if not isinstance(sheet_date, datetime):
            raise ValueError(f'Data inválida na célula: {sheet_date}')

    except IndexError:
        raise ValueError(f'Índice {month + MONTH_DESLOCATE} fora do intervalo de `headers`.')

    if month != sheet_date.month or year != sheet_date.year:
        raise ValueError('Mês e ano não batem com célula na planilha.')
