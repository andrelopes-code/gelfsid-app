from datetime import datetime

import openpyxl
import rich
from rich import print

from gelfmp.management.commands.collectors.types import Aliases
from gelfmp.models import CharcoalMonthlyPlan, Supplier

from .utils import get_best_matches

CHARCOAL_SCHED_PATH = r'.sample-data\schedule_2024.xlsx'

GELF_ID = 1

MIN_COL = 2
MAX_COL = 19
MIN_ROW = 7
MONTH_DESLOC = 3  # Meses começam na quarta coluna.
NAME_COL = 0
FARM_COL = 1
MINIMUM_SIMILARITY = 95

MONTHS = {
    datetime(2024, 1, 1, 0, 0),
    datetime(2024, 2, 1, 0, 0),
    datetime(2024, 3, 1, 0, 0),
    datetime(2024, 4, 1, 0, 0),
    datetime(2024, 5, 1, 0, 0),
    datetime(2024, 6, 1, 0, 0),
    datetime(2024, 7, 1, 0, 0),
    datetime(2024, 8, 1, 0, 0),
    datetime(2024, 9, 1, 0, 0),
    datetime(2024, 10, 1, 0, 0),
    datetime(2024, 11, 1, 0, 0),
    datetime(2024, 12, 1, 0, 0),
}


def gelf_schedule(rows, month, year):
    _ = next(rows)
    programmed_volume_row = next(rows)
    programmed_volume = programmed_volume_row[month]

    gelf = Supplier.objects.get(id=GELF_ID)
    if 'GELF' not in gelf.corporate_name.upper():
        raise ValueError(f'Esperado fornecedor GELF com id {GELF_ID}, encontrado: {gelf}')

    CharcoalMonthlyPlan.objects.create(
        supplier=gelf,
        month=month - MONTH_DESLOC,
        year=year,
        programmed_volume=programmed_volume,
    )


def validate_month(headers, month):
    sheet_month = headers[month + MONTH_DESLOC]

    if sheet_month not in MONTHS:
        raise ValueError('Mês inválido ou formato da planilha inválido.')
    if not month == sheet_month.month:
        raise ValueError('Mês não bate com célula na planilha.')

    return month + MONTH_DESLOC


def collect():
    documents_workbook = openpyxl.load_workbook(CHARCOAL_SCHED_PATH, data_only=True)
    documents_sheet = documents_workbook.active
    month = 9
    year = 2024

    if not documents_sheet:
        raise ValueError(f'document `{CHARCOAL_SCHED_PATH}` does not have an active sheet')

    headers = next(
        documents_sheet.iter_rows(
            values_only=True,
            min_col=MIN_COL,
            max_col=MAX_COL,
            min_row=MIN_ROW - 1,
        )
    )
    month = validate_month(headers, month)

    rows = documents_sheet.iter_rows(values_only=True, min_col=MIN_COL, max_col=MAX_COL, min_row=MIN_ROW + 4)

    # gelf_schedule(rows, month, year)

    r1 = next(rows)
    r2 = next(rows)

    while r1[0] and r1[0] != '-':
        print(r1[0])
        r1 = next(rows)
        r2 = next(rows)


# def collect():
#     suppliers_names = list(Supplier.objects.values_list('corporate_name', flat=True))
#     if not suppliers_names:
#         raise ValueError('Não existem fornecedores no banco de dados')

#     aliases = Aliases('charcoal.alias.json')

#     for entry_supplier_name, group in entries.groupby('FORNECEDOR_E_FAZENDA'):
#         if entry_supplier_name in aliases:
#             process_data(aliases.get(entry_supplier_name), group)
#             continue

#         # Encontra as melhores correspondências de fornecedor
#         best_matches = get_best_matches(entry_supplier_name, suppliers_names)

#         # Pega a melhor correspondência (a primeira)
#         similarity, supplier_name = best_matches[0]

#         if similarity >= MINIMUM_SIMILARITY:
#             process_data(supplier_name, group)
#             continue

#         rich.print(f'[red]FORNECEDOR NÃO ENCONTRADO: {entry_supplier_name}[/red]')
#         for i, (sim, name) in enumerate(best_matches):
#             rich.print(f'  {i + 1} => {name} [{sim:.2f}]')

#         option = Prompt.ask(
#             '[green]O nome não encontrado corresponde a alguma dessas opções? '
#             '(digite o número ou pressione enter para pular)[/green]',
#         )
#         print()

#         if option.strip().isdigit():
#             option = int(option.strip()) - 1
#             valid_index = 0 <= option < len(best_matches)

#             if valid_index:
#                 supplier_name = best_matches[option][1]
#                 aliases.add(entry_supplier_name, supplier_name)
#                 process_data(supplier_name, group)
